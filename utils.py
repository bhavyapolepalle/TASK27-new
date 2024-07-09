import openpyxl

def read_test_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_data.append(row)
    return test_data

def write_test_result(file_path, row, result):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.cell(row=row, column=7, value=result)
    workbook.save(file_path)